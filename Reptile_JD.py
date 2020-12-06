'''
FilePath: \Source\Reptile_JD.py
Author: Eventi
Date: 2020-12-05 00:15:11
LastEditors: Eventi
LastEditTime: 2020-12-06 13:03:32
Description:
'''
from logging import PercentStyle
from selenium import webdriver
import time
from bs4 import BeautifulSoup
import openpyxl
from selenium.webdriver.remote.switch_to import SwitchTo
import tkinter

LINE_MAX_NUM = 1000000  # 数据最大条数
data_line = dict()  # 数据字典
html_urls = []  # 获取的商品详情页链接列表

'''
description: 通过向下滑动,加载整个页面的完整数据
param {*} browser 浏览器对象
return {*}
'''
def buffer(browser):
    for i in range(1000):
        time.sleep(0.001)
        browser.execute_script('window.scrollBy(0,100)', '')

'''
description: 将数据保存到excell中
param {*}
return {*}
'''
def data_save(line):
    # 输入到Excel表格中
    wb = openpyxl.Workbook()
    ws = wb.active
    for xl_line in range(0, line):
        count = 1
        xlsx_line = 0x41
        xlsx_str_start = ''
        for key in data_line.keys():
            if 1 == count % 26 and 1 != count:
                xlsx_str_start = chr(0x40+(count//26))
                xlsx_line = 0x41
            else:
                pass
            if 0 == xl_line:
                xlsx_str = xlsx_str_start + chr(xlsx_line) + '1'
                ws[xlsx_str] = key
            xlsx_str = xlsx_str_start + chr(xlsx_line) + str(xl_line+2)
            ws[xlsx_str] = data_line[key][xl_line]
            xlsx_line = xlsx_line + 1
            count = count+1
    wb.save('../data/data.xlsx')

'''
description: 根据搜索结果的URL地址获取整个搜索结果商品的详情数据
param {*} search_url 传入搜索结果的url地址
return {*}
'''
def getData(search_url):
    browser = webdriver.Edge(r'c:/msedgedriver.exe')  # 使用浏览器驱动启动浏览器
    print(search_url[8:15])
    # if('search'==search_url[8:13]):#判断是否为全网搜索,如果为全网搜索则执行一下语句,否则为店内搜索,执行else以下的内容
    if 's' == (search_url[8]) and 'e' == search_url[9] and 'a' == search_url[10] and 'r' == search_url[11] and 'c' == search_url[12] and 'h' == search_url[13] and '.' == search_url[14]:
        browser.get(search_url)  # 打开搜索结果页面
        while True:  # 当搜索结果结果有多页时,通过点击下一页访问后面的页面,直至下一页内容无效则退出此循环
            buffer(browser)  # 通过向下滑动来加载整个页面
            divs = browser.find_elements_by_class_name('gl-i-wrap')
            for div in divs:
                detail_url = div.find_element_by_class_name(
                    'p-img').find_element_by_tag_name('a').get_attribute('href')
                # if "http" == detail_url[0:4]:#判断href中是否是http地址,是则加入到详情页地址列表中
                if 'h' == detail_url[0] and 'p' == detail_url[3]:
                    html_urls.append(detail_url)
            try:
                nexe_page = browser.find_element_by_class_name(
                    'pn-next')  # 如果页面中无下一页按钮则通过抛出异常来退出循环
            except:
                break
            if(None == nexe_page.get_attribute('href')):  # 如果是最后一个页面,则通过检查href属性为空来退出循环
                break
            else:
                nexe_page.click()
    else:
        browser.get(search_url)  # 打开搜索结果页面
        while True:  # 当搜索结果结果有多页时,通过点击下一页访问后面的页面,直至下一页内容无效则退出此循环
            buffer(browser)  # 通过向下滑动来加载整个页面
            lis = browser.find_element_by_class_name('jSearchListArea').find_element_by_class_name(
                'j-module').find_elements_by_tag_name('li')
            for li in lis:
                detail_url = li.find_element_by_tag_name(
                    'a').get_attribute('href')
                # if "http" == detail_url[0:4]:#判断href中是否是http地址,是则加入到详情页地址列表中
                if 'h' == detail_url[0] and 'p' == detail_url[3]:
                    html_urls.append(detail_url)
            try:
                nexe_page = browser.find_element_by_class_name(
                    'jPage')  # 如果页面中无下一页按钮则通过抛出异常来退出循环
            except:
                break
            next_as = nexe_page.find_elements_by_tag_name('a')
            for next_a in next_as:
                # 通过a标签内文本和href属性来检查是否有下一页按钮,
                if '下一页' == next_a.text and None != next_a.get_attribute('href'):
                    browser.get(next_a.get_attribute('href'))
                    continue
            break

    try:
        line = 0  # 当前获取的数据索引值 初始化为0从头获取详情页数据,并将数据标题放到字典的key中将数据内容通过列表的方式放到字典的value中
        for html_url in html_urls:
            browser.get(html_url)  # 打开详情页面
            buffer(browser)  # 通过向下滑动来加载整个页面(主要是为了加载出评论)
            html = browser.find_element_by_tag_name(
                'html').get_attribute('outerHTML')
            soup = BeautifulSoup(html, "html.parser")  # 部分数据通过BeautifulSoup解析

            # 详情 每个key都通过异常捕获的方式来判断是否创建了这个key,每个vulue都通过异常捕获,如果异常则value赋值为空
            try:
                data_line["抓取时间"]
            except:
                data_line["抓取时间"] = [None]*LINE_MAX_NUM
            data_line["抓取时间"][line] = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime())
            try:
                data_line["京东链接"]
            except:
                data_line["京东链接"] = [None]*LINE_MAX_NUM
            data_line["京东链接"][line] = html_url
            try:
                data_line["店铺名称"]
            except:
                data_line["店铺名称"] = [None]*LINE_MAX_NUM
            try:  # 店铺名称在一些商品页无存在,则会抛出异常
                data_line["店铺名称"][line] = soup.find(
                    class_='J-hove-wrap').div.div.a.string
            except:
                data_line["店铺名称"][line] = ''
            try:
                data_line["商品标题"]
            except:
                data_line["商品标题"] = [None]*LINE_MAX_NUM
            try:
                data_line["商品标题"][line] = browser.find_element_by_class_name(
                    'sku-name').text
            except:
                data_line["商品标题"][line] = ''
            try:
                data_line["京东价"]
            except:
                data_line["京东价"] = [None]*LINE_MAX_NUM
            try:
                data_line["京东价"][line] = browser.find_element_by_class_name(
                    'p-price').find_elements_by_xpath('span')[1].text
            except:
                data_line["京东价"][line] = ''
            original_price = soup.find(id='page_hx_price')
            try:
                data_line["原价"]
            except:
                data_line["原价"] = [None]*LINE_MAX_NUM
            if None != original_price:
                data_line["原价"][line] = original_price.string.replace(  # 去掉人民币符号
                    '￥', '', 1)
            else:
                data_line["原价"][line] = ''

            # 好评度
            try:  # 好评度需要向下滑动才能加载到,如果未加载出来则抛出异常,在异常中多次尝试滑动
                percent = browser.find_element_by_class_name('comment-percent')
            except:
                buffer(browser)
                for times in range(0, 30):  # 最多尝试30次
                    try:
                        percent = browser.find_element_by_class_name(
                            'comment-percent')
                    except:
                        buffer(browser)
                        continue
                    break
            percent = browser.find_element_by_class_name('comment-percent')
            key = percent.find_element_by_tag_name('strong').text
            try:
                data_line[key]
            except:
                data_line[key] = [None]*LINE_MAX_NUM
            try:
                data_line[key][line] = percent.find_element_by_tag_name(
                    'div').text
            except:
                data_line[key][line] = ''

            # 评论
            comments = browser.find_element_by_class_name(
                'filter-list').find_elements_by_tag_name('li')
            for comment in comments:
                try:
                    a = comment.find_element_by_tag_name('a')
                except:
                    pass
                else:
                    value = a.find_element_by_tag_name('em').text  # 获取vulue值
                    key = a.text.replace(value, '', 1)  # 获取key值
                    value = value.replace('\x28', '', 1)  # 去掉左括号
                    value = value.replace('\x29', '', 1)  # 去掉右括号
                    if '' != key:
                        try:
                            data_line[key]
                        except:
                            data_line[key] = [None]*LINE_MAX_NUM
                        data_line[key][line] = value

            # 包装
            packing = soup.find(class_='package-list')
            key = packing.h3.string
            try:
                data_line[key]
            except:
                data_line[key] = [None]*LINE_MAX_NUM
            data_line[key][line] = packing.p.string

            # 规格
            specifications = soup.find_all(class_='Ptable-item')
            for each_specifications in specifications:
                key_h3 = each_specifications.h3.string+'·'  # 获取规范类别
                dl = each_specifications.dl.find_all(class_='clearfix')
                for each_dl in dl:  # 获取类别中每一项内容
                    key = key_h3+each_dl.dt.string
                    dd = each_dl.find_all('dd')
                    if (1 == len(dd)):
                        value = dd[0].string
                    else:
                        value = dd[1].string
                    try:
                        data_line[key]
                    except:
                        data_line[key] = [None]*LINE_MAX_NUM
                    data_line[key][line] = value
            line = line + 1  # 此页面数据获取完成,+1获取下一商品数据
    finally:  # 当以上任何一处出现无处理的异常,都要要执行下面的内容,将已获取的数据进行保存
        data_save(line)  # 保存数据
    browser.quit()  # 关闭浏览器

top = tkinter.Tk()#获取宽口对象
top.title('爬取数据')#设置串口标题
L1 = tkinter.Label(top, text="网址")
L1.pack(side=tkinter.LEFT)#设置label
E1 = tkinter.Entry(top, bd=5)#设置输入框
E1.pack(side=tkinter.LEFT)
def buttonCommand():#按键处理函数
    getData(E1.get())#每次按下则执行获取数据函数
B1 = tkinter.Button(top, bd=3, text='确定', command=buttonCommand)#设置按钮
B1.pack(side=tkinter.LEFT)
top.mainloop()#启动界面
